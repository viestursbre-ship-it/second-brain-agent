"""
Second Brain Agent - Multi-format Document Scanner (.txt, .md, .pdf, .docx, .xlsx)
"""

from pathlib import Path
import docx
from openpyxl import load_workbook
from pypdf import PdfReader


class DocumentScanner:

  @staticmethod
  def extract_text_from_file(uploaded_file) -> str:
    filename = uploaded_file.name.lower()

    # 1. TXT / MD
    if filename.endswith((".txt", ".md")):
      return uploaded_file.read().decode("utf-8", errors="ignore")

    # 2. PDF
    elif filename.endswith(".pdf"):
      reader = PdfReader(uploaded_file)
      return "\n".join([page.extract_text() or "" for page in reader.pages])

    # 3. Word (.docx)
    elif filename.endswith(".docx"):
      doc = docx.Document(uploaded_file)
      return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

    # 4. Excel (.xlsx)
    elif filename.endswith(".xlsx"):
      wb = load_workbook(uploaded_file, data_only=True)
      lines = []
      for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
          row_vals = [
              str(val) for val in row if val is not None and str(val).strip()
          ]
          if row_vals:
            lines.append(" | ".join(row_vals))
      return "\n".join(lines)

    return ""
